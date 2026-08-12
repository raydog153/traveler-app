"""One-time extraction of the Gas and Bus Maintenance tabs from a downloaded
snapshot of the "Bus Living - Our Spot" Google Sheet into committed JSON
fixtures (fixtures/gas_raw.json, fixtures/maint_raw.json).

Run once against a manually-downloaded XLSX export of the sheet:

    python -m seed.extract_gas_maintenance /path/to/bus_living.xlsx

Rationale for using the XLSX export + openpyxl rather than Drive's plain-text
export: the text export silently truncates the Gas tab partway through 2025
(dropping ~45 rows with no error), while the XLSX export via openpyxl reads
every row faithfully.

Row/column layout (verified by inspection):
  - `Gas` tab: header row 1, data rows starting at row 2. Columns A-I are
    Date, Milage, Gallons, Price, Cost/gallon, Driven Milage, MPG, Notes,
    City. Cost/gallon, Driven Milage, and MPG are precomputed in the sheet
    and are re-derived by the app instead of being imported. Columns J-P are
    an unrelated per-year sidebar summary, ignored. Data rows are identified
    by column A holding an actual datetime (this also naturally skips the
    trailing blank rows and the "Total" summary row).
  - `Bus Maintenance` tab: header row 1, data rows 2-76. Columns A-F are
    Date, Expense, Place, Milage, Place (second "Place" column is actually
    the vendor), Cost. Row 77 is a "Total" row. Rows after ~77 contain
    unrelated trackers (last oil change, bus weight log, shocks log) that
    also have datetime-typed dates in column A, so — unlike the Gas tab —
    extraction here is bounded to rows 2-76 explicitly rather than relying
    on the datetime-type check alone.

Known source-data correction (manually applied to fixtures/gas_raw.json,
NOT automatically re-applied by this script): the "Pembroke, GA" row
(odometer 121657, 37.528 gal, $130.82) is dated 2022-12-28 in the sheet, but
this is a year typo -- it should be 2021-12-28. Evidence: its $3.49/gal cost
matches early-2022/late-2021 fuel prices, not the $4.10-4.60/gal seen on the
surrounding late-Dec-2022 rows; its sheet-precomputed "driven" value (209 mi)
was computed against the row physically above it in the sheet (2021-12-05),
not against the true Dec-2022 predecessor (2022-12-27); and the corrected
date makes the 2021-12-05 -> 2021-12-28 -> 2022-01-09 odometer sequence
consistent with each row's own precomputed driven/mpg values. Taken at face
value, the typo'd date sorts this row into the wrong chronological position
and produces a ~15,000-mile phantom jump into the following (2023-02-02)
fill-up's computed mileage/MPG. If this script is re-run against a fresh
sheet export, check whether this row's date has since been corrected
upstream in the sheet before re-applying this fix.
"""

import datetime
import json
import sys
from pathlib import Path

import openpyxl

FIXTURES_DIR = Path(__file__).parent / "fixtures"

GAS_MAX_ROW = 205
MAINTENANCE_MAX_ROW = 76


def _iso(d: datetime.datetime) -> str:
    return d.date().isoformat()


def extract_gas(ws) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=GAS_MAX_ROW, values_only=True):
        date_cell = row[0]
        if not isinstance(date_cell, datetime.datetime):
            continue
        _date, milage, gallons, price, _cpg, _driven, _mpg, notes, city = row[:9]
        rows.append(
            {
                "date": _iso(date_cell),
                "odometer_miles": float(milage),
                "gallons": float(gallons),
                "price": float(price),
                "notes": (notes or "").strip(),
                "city": (city or "").strip(),
            }
        )
    return rows


def extract_maintenance(ws) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=MAINTENANCE_MAX_ROW, values_only=True):
        date_cell = row[0]
        if not isinstance(date_cell, datetime.datetime):
            continue
        _date, expense, place, milage, vendor, cost = row[:6]
        rows.append(
            {
                "date": _iso(date_cell),
                "expense": (expense or "").strip(),
                "place": (place or "").strip(),
                "odometer_miles": float(milage) if milage is not None else None,
                "vendor": (vendor or "").strip(),
                "cost": float(cost),
            }
        )
    return rows


def main(xlsx_path: str) -> None:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    gas_rows = extract_gas(wb["Gas"])
    maint_rows = extract_maintenance(wb["Bus Maintenance"])

    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    (FIXTURES_DIR / "gas_raw.json").write_text(json.dumps(gas_rows, indent=2))
    (FIXTURES_DIR / "maint_raw.json").write_text(json.dumps(maint_rows, indent=2))

    print(f"gas: {len(gas_rows)} rows -> {FIXTURES_DIR / 'gas_raw.json'}")
    print(f"maintenance: {len(maint_rows)} rows -> {FIXTURES_DIR / 'maint_raw.json'}")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("usage: python -m seed.extract_gas_maintenance /path/to/bus_living.xlsx")
        sys.exit(1)
    main(sys.argv[1])
